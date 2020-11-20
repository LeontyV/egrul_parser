import openpyxl


def to_file(folder_name, row_info):
    file_name = f'{folder_name}\\to_reestr.xlsx'
    file_name = file_name.replace('\\\\', '\\')
    try:
        wb = openpyxl.load_workbook(filename=file_name)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        wb.save(file_name)
        wb = openpyxl.load_workbook(filename=file_name)

    count = 1
    if (wb):
        try:
            ws_ul = wb["ЮЛ"]
            ws_fl = wb["ФЛ"]
        except KeyError:
            ws_ul = wb.create_sheet("ЮЛ", 0)
            ws_fl = wb.create_sheet("ФЛ", 1)
    else:
        wb = openpyxl.Workbook()
        ws_ul = wb.create_sheet("ЮЛ", 0)
        ws_fl = wb.create_sheet("ФЛ", 1)

    # создаем заголовки для ЮЛ
    ws_ul.cell(row=1, column=1, value='Серийный номер')
    ws_ul.cell(row=1, column=2, value='начало')
    ws_ul.cell(row=1, column=3, value='конец')
    ws_ul.cell(row=1, column=4, value='Общее имя')
    ws_ul.cell(row=1, column=5, value='ФИО')
    ws_ul.cell(row=1, column=6, value='ИНН')
    ws_ul.cell(row=1, column=7, value='Организация')
    ws_ul.cell(row=1, column=8, value='Адрес')
    ws_ul.cell(row=1, column=9, value='ОГРН')
    ws_ul.cell(row=1, column=10, value='реквизиты свидетельства о государственной регистрации юридического лица')
    ws_ul.cell(row=1, column=11, value='реквизиты документов о государственной регистрации юридического лица в соответствии с законодательством иностранного государства')
    ws_ul.cell(row=1, column=12, value='основные реквизиты (наименование, номер и дата выдачи) документа, подтверждающего право лица, выступавшего от имени заявителя, обращаться за получением квалифицированного сертификата')

    # создаем заголовки для ФЛ
    ws_fl.cell(row=1, column=1, value='Серийный номер')
    ws_fl.cell(row=1, column=2, value='начало')
    ws_fl.cell(row=1, column=3, value='конец')
    ws_fl.cell(row=1, column=4, value='Общее имя')
    ws_fl.cell(row=1, column=6, value='ИНН')
    ws_fl.cell(row=1, column=9, value='СНИЛС')

    max_row_ul = ws_ul.max_row
    count_ul = max_row_ul + 1
    max_row_fl = ws_fl.max_row
    count_fl = max_row_fl + 1
    if row_info.get('Организация'):
        ws_ul.cell(row=count_ul, column=1, value=row_info['Серийный номер'])
        ws_ul.cell(row=count_ul, column=2, value=row_info['начало'])
        ws_ul.cell(row=count_ul, column=3, value=row_info['конец'])
        ws_ul.cell(row=count_ul, column=4, value=row_info['Общее имя'])
        if row_info['ФИО'] != '':
            ws_ul.cell(row=count_ul, column=5, value=row_info['ФИО'])
        else:
            ws_ul.cell(row=count_ul, column=5, value=row_info['Организация'])
        ws_ul.cell(row=count_ul, column=6, value=row_info['ИНН'])
        ws_ul.cell(row=count_ul, column=7, value=row_info['Организация'])
        ws_ul.cell(row=count_ul, column=8, value=row_info['Адрес'])
        ws_ul.cell(row=count_ul, column=9, value=row_info['ОГРН'])
    else:
        ws_fl.cell(row=count_fl, column=1, value=row_info['Серийный номер'])
        ws_fl.cell(row=count_fl, column=2, value=row_info['начало'])
        ws_fl.cell(row=count_fl, column=3, value=row_info['конец'])
        ws_fl.cell(row=count_fl, column=4, value=row_info['Общее имя'])
        ws_fl.cell(row=count_fl, column=5, value=row_info['ИНН'])
        ws_fl.cell(row=count_fl, column=6, value=row_info['СНИЛС'])

    wb.save(file_name)